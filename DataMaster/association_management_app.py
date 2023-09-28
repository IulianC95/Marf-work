import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
import json
import datetime
import os
import pygetwindow as gw
import re

def initialize_association_files():
    for assoc in associations:
        json_file_path = os.path.join("json", f"{assoc}_months.json")
        
        # Dacă fișierul nu există, îl creăm cu valori inițiale
        if not os.path.exists(json_file_path):
            initial_data = [datetime.datetime.now().strftime("%B %Y")]
            save_to_json(f"{assoc}_months.json", initial_data)



def export_association_status():
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # Initialize DataFrame with a large number of columns for status
    max_columns = 20  # Assuming a maximum of 20 unchecked items
    columns = ["Asociatie"] + [f"Status_{i+1}" for i in range(max_columns)]
    df = pd.DataFrame(columns=columns)

    for assoc in associations:
        month_list = load_from_json(f"{assoc}_months.json") or []
        suppliers = load_from_json(f"{assoc}_{month_list[-1] if month_list else 'N/A'}_suppliers") or {}
        notes = load_from_json(f"{assoc}_{month_list[-1] if month_list else 'N/A'}_notes.json") or {}
        
        items = {**suppliers, **notes}  # Combining suppliers and notes

        row_data = {"Asociatie": assoc}
        if month_list:
            last_month = month_list[-1]

            if all(items.values()):
                row_data["Status_1"] = "Se poate închide luna"
            else:
                unbifated = [f"{k} (F)" if k in suppliers else f"{k} (N)" for k, v in items.items() if not v]
                for i, item in enumerate(unbifated):
                    row_data[f"Status_{i+1}"] = item

        else:
            row_data["Status_1"] = "Nicio luna disponibila"

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True, sort=False)

    # Save DataFrame to an Excel file
    excel_path = "Status_Asociatii.xlsx"
    df.to_excel(excel_path, index=False)

    # Open the Excel file to adjust the column widths
    wb = load_workbook(excel_path)
    ws = wb.active  # Get the active sheet

    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_path)



def get_open_association():
    titles = gw.getAllTitles()
    for title in titles:
        if 'BlocManagerNET' in title:
            match = re.search(r"Asociatia de proprietari (.+?) --", title, re.IGNORECASE)
            if match:
                return match.group(1).lower().strip()
    return None

def sanitized_string(s):
    return ''.join(e for e in s if e.isalnum()).lower()

def find_matching_association(association_name):
    # Deschide fisierul JSON
    with open('json\\associations.json', 'r') as f:
        json_data = json.load(f)

    sanitized_association_name = sanitized_string(association_name)
    sanitized_json_data = [sanitized_string(x) for x in json_data]

    print(f"Sanitized Detected: '{sanitized_association_name}'")  # Pentru debug
    print(f"Sanitized Available: {sanitized_json_data}")    # Pentru debug

    if sanitized_association_name in sanitized_json_data:
        index = sanitized_json_data.index(sanitized_association_name)
        return json_data[index]
    return None


def detect_bloc_manager_association():
    global last_association

    open_association = get_open_association()
    
    if open_association and (last_association is None or last_association != open_association):
        print(f"Asociatia detectată: {open_association}")  # Pentru debug
        matching_association = find_matching_association(open_association)
        
        if matching_association:
            selected_association.set(matching_association)
            
        update_association_title()
        refresh_suppliers()
        
        last_association = open_association  # Actualizează last_association cu asociatia curentă

    root.after(5000, detect_bloc_manager_association)



def save_to_json(filename, data):
    with open(os.path.join("json", filename), 'w') as f:
        json.dump(data, f)

def load_from_json(filename):
    try:
        with open(os.path.join("json", filename), 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return None

closed_months = load_from_json("closed_months.json") or {}
last_association = None

def update_association_title(*args):
    association_title.config(text=selected_association.get())
    month_list = load_from_json(f"{selected_association.get()}_months.json") or [datetime.datetime.now().strftime("%B %Y")]
    month_dropdown['values'] = month_list
    selected_month.set(month_list[-1])
    refresh_suppliers()
    refresh_notes()

def add_association():
    new_association = simpledialog.askstring("Input", "Introdu numele asociatiei:")
    if new_association:
        associations.append(new_association)
        selected_association.set(new_association)
        save_to_json("associations.json", associations)
        association_dropdown['values'] = associations

def add_supplier():
    new_window = tk.Toplevel(root)
    new_window.title('Adaugă furnizor')
    new_window.geometry('400x100')

    tk.Label(new_window, text='Introduce numele furnizorului:').pack(pady=10)
    supplier_name = tk.Entry(new_window, width=25)
    supplier_name.pack(fill=tk.X, padx=10)

    def on_ok():
        new_supplier = supplier_name.get()
        if new_supplier:
            suppliers = load_from_json(f"{selected_association.get()}_{selected_month.get()}_suppliers") or {}
            suppliers[new_supplier] = False
            save_to_json(f"{selected_association.get()}_{selected_month.get()}_suppliers", suppliers)
            refresh_suppliers()
        new_window.destroy()

    tk.Button(new_window, text='OK', command=on_ok).pack(pady=5)



def close_month():
    password = simpledialog.askstring("Password", "Introduceți parola:")
    if password != "1234":
        return

    if selected_association.get() not in closed_months:
        closed_months[selected_association.get()] = []

    closed_months[selected_association.get()].append(selected_month.get())
    save_to_json("closed_months.json", closed_months)

    save_checkboxes()

    current_month_str = selected_month.get()
    current_month = datetime.datetime.strptime(current_month_str, "%B %Y")
    next_month = current_month + datetime.timedelta(days=32)
    next_month_str = next_month.strftime("%B %Y")
    month_list = load_from_json(f"{selected_association.get()}_months.json") or [datetime.datetime.now().strftime("%B %Y")]
    month_list.append(next_month_str)
    month_dropdown['values'] = month_list
    selected_month.set(next_month_str)
    save_to_json(f"{selected_association.get()}_months.json", month_list)

    suppliers = load_from_json(f"{selected_association.get()}_{current_month_str}_suppliers") or {}
    new_suppliers = {k: False for k in suppliers.keys()}
    save_to_json(f"{selected_association.get()}_{next_month_str}_suppliers", new_suppliers)

    refresh_suppliers()
    refresh_notes()

def refresh_suppliers():
    for widget in items_frame.winfo_children():
        if getattr(widget, "tag", None) == "supplier":
            widget.destroy()

    suppliers = load_from_json(f"{selected_association.get()}_{selected_month.get()}_suppliers") or {}
    is_month_closed = selected_month.get() in closed_months.get(selected_association.get(), [])

    if suppliers:
        for supplier, checked in suppliers.items():
            var = tk.BooleanVar(value=suppliers[supplier] if supplier in suppliers else False)
            cb = tk.Checkbutton(items_frame, text=supplier, variable=var, state='disabled' if is_month_closed else 'normal', command=check_all_checked)
            cb.var = var
            cb.pack()
            cb.tag = "supplier"
            cb.bind("<ButtonRelease-1>", lambda e: check_all_checked())
            check_all_checked()


def add_note():
    new_window = tk.Toplevel(root)
    new_window.title('Adaugă notiță')
    new_window.geometry('400x100')

    tk.Label(new_window, text='Scrie notița:').pack(pady=10)
    note_text = tk.Entry(new_window)
    note_text.pack(fill=tk.X, padx=10)

    def on_ok():
        new_note = note_text.get()
        if new_note:
            notes = load_from_json(f"{selected_association.get()}_{selected_month.get()}_notes.json") or {}
            notes[new_note] = False
            save_to_json(f"{selected_association.get()}_{selected_month.get()}_notes.json", notes)
            refresh_notes()
        new_window.destroy()

    tk.Button(new_window, text='OK', command=on_ok).pack(pady=5)

def check_all_checked():
    all_checked = all(widget.var.get() for widget in items_frame.winfo_children() if isinstance(widget, tk.Checkbutton))
    if all_checked:
        items_frame.config(bg='#7FDD7D')
    else:
        items_frame.config(bg='white')


def refresh_notes():
    for widget in items_frame.winfo_children():
         if getattr(widget, "tag", None) == "note":
             widget.destroy()

    notes = load_from_json(f"{selected_association.get()}_{selected_month.get()}_notes.json") or {}
    is_month_closed = selected_month.get() in closed_months.get(selected_association.get(), [])

    if notes:
        for note, checked in notes.items():
            var = tk.BooleanVar(value=notes[note] if note in notes else False)
            cb = tk.Checkbutton(items_frame, text=note, variable=var, state='disabled' if is_month_closed else 'normal', command=check_all_checked)
            cb.var = var
            cb.pack()
            cb.tag = "note"
            cb.bind("<ButtonRelease-1>", lambda e: check_all_checked())
            check_all_checked()

def save_checkboxes():
    suppliers = {}
    notes = {}
    for widget in items_frame.winfo_children():
        if isinstance(widget, tk.Checkbutton):
            if getattr(widget, "tag", None) == "supplier":
                suppliers[widget.cget("text")] = widget.var.get()
            elif getattr(widget, "tag", None) == "note":
                notes[widget.cget("text")] = widget.var.get()
    save_to_json(f"{selected_association.get()}_{selected_month.get()}_suppliers", suppliers)
    save_to_json(f"{selected_association.get()}_{selected_month.get()}_notes.json", notes)

def delete_note():
    delete_window = tk.Toplevel(root)
    delete_window.title("Șterge Notiță")

    notes = load_from_json(f"{selected_association.get()}_{selected_month.get()}_notes.json") or {}
    
    # Lista pentru a ține notițele selectate pentru ștergere
    selected_notes = []
    
    for note in notes.keys():
        var = tk.BooleanVar()
        cb = tk.Checkbutton(delete_window, text=note, variable=var)
        cb.var = var
        cb.pack()
        
        selected_notes.append((note, var))
    
    # Buton pentru a confirma ștergerea
    tk.Button(delete_window, text="Șterge", command=lambda: confirm_delete_notes(selected_notes)).pack()

def confirm_delete_notes(selected_notes):
    notes = load_from_json(f"{selected_association.get()}_{selected_month.get()}_notes.json") or {}
    
    for note, var in selected_notes:
        if var.get():
            del notes[note]
    
    save_to_json(f"{selected_association.get()}_{selected_month.get()}_notes.json", notes)
    refresh_notes()

def update_button_states():
    is_month_closed = selected_month.get() in closed_months.get(selected_association.get(), [])
    state = tk.DISABLED if is_month_closed else tk.NORMAL
    add_supplier_button["state"] = state
    add_note_button["state"] = state
    save_button["state"] = state
    close_month_button["state"] = state
    delete_supplier_button["state"] = state
    delete_note_button["state"] = state

def delete_supplier():
    delete_window = tk.Toplevel(root)
    delete_window.title("Șterge Furnizor")

    suppliers = load_from_json(f"{selected_association.get()}_{selected_month.get()}_suppliers") or {}
    
    # Lista pentru a ține furnizorii selectați pentru ștergere
    selected_suppliers = []
    
    for supplier in suppliers.keys():
        var = tk.BooleanVar()
        cb = tk.Checkbutton(delete_window, text=supplier, variable=var)
        cb.var = var
        cb.pack()
        
        selected_suppliers.append((supplier, var))
    
    # Buton pentru a confirma ștergerea
    tk.Button(delete_window, text="Șterge", command=lambda: confirm_delete(selected_suppliers)).pack()




root = tk.Tk()
root.title("Data Master v1.02")
root.geometry("1200x800")

left_frame = tk.Frame(root, bg='#52817F')
left_frame.pack(side=tk.LEFT, fill=tk.Y)
left_frame.pack_propagate(0)
left_frame.config(width=250)

associations = load_from_json("associations.json") or []
initialize_association_files()
selected_association = tk.StringVar()
if associations:
    selected_association.set(associations[0])
selected_association.trace_add("write", lambda *args: [update_association_title(), refresh_suppliers(), refresh_notes()])
association_dropdown = ttk.Combobox(left_frame, textvariable=selected_association, values=associations)
association_dropdown.pack(pady=10, fill=tk.X)
selected_association.trace_add("write", update_association_title)

initial_association = associations[0] if associations else None
month_list = load_from_json(f"{initial_association}_months.json") or [datetime.datetime.now().strftime("%B %Y")]
selected_month = tk.StringVar()
selected_month.set(month_list[-1])
selected_month.trace_add("write", lambda *args: [refresh_suppliers(), refresh_notes()])
month_dropdown = ttk.Combobox(left_frame, textvariable=selected_month, values=month_list)
month_dropdown.pack(pady=10, fill=tk.X)

tk.Button(left_frame, text="Asociatie noua", command=add_association).pack(pady=5)
tk.Button(left_frame, text="Status Asociatii", command=export_association_status).pack(pady=5)


right_frame = tk.Frame(root, width=1080, bg='#52817F')
right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=1)

association_title = tk.Label(right_frame, text="Asociatie", font=("Arial", 24), bg='#52817F')
association_title.pack(pady=10)

suppliers_label = tk.Label(right_frame, text="Furnizori / Notite", font=("Arial", 20), bg='#52817F')
suppliers_label.pack(pady=10)
button_frame = tk.Frame(right_frame, bg='#52817F')
button_frame.pack(pady=5)

add_supplier_button = tk.Button(button_frame, text="Adauga furnizor", command=add_supplier)
add_supplier_button.grid(row=0, column=0, padx=5)

add_note_button = tk.Button(button_frame, text="Adauga notita", command=add_note)
add_note_button.grid(row=0, column=1, padx=5)

delete_supplier_button = tk.Button(button_frame, text="Șterge Furnizor", command=delete_supplier)
delete_supplier_button.grid(row=1, column=0, padx=5, pady=5)

delete_note_button = tk.Button(button_frame, text="Șterge Notiță", command=delete_note)
delete_note_button.grid(row=1, column=1, padx=5, pady=5)


items_frame = tk.Frame(right_frame, bg='white')
items_frame.pack(fill=tk.BOTH, expand=1)
save_button = tk.Button(right_frame, text="Save", command=save_checkboxes)
save_button.pack(pady=5)




close_month_button = tk.Button(right_frame, text="Închide luna", command=close_month)
close_month_button.pack(side=tk.LEFT, padx=5)

selected_month.trace_add("write", lambda *args: [refresh_suppliers(), refresh_notes(), update_button_states()])
selected_association.trace_add("write", lambda *args: [update_association_title(), update_button_states()])


refresh_notes()
refresh_suppliers()

root.after(5000, detect_bloc_manager_association)



def confirm_delete(selected_suppliers):
    suppliers = load_from_json(f"{selected_association.get()}_{selected_month.get()}_suppliers") or {}
    
    for supplier, var in selected_suppliers:
        if var.get():
            del suppliers[supplier]
    
    save_to_json(f"{selected_association.get()}_{selected_month.get()}_suppliers", suppliers)
    refresh_suppliers()

from tkinter import PhotoImage

def toggle_main_window():
    if root.winfo_viewable():
        root.withdraw()
    else:
        root.deiconify()

icon_root = tk.Toplevel(root)
icon_root.overrideredirect(1) 
icon_root.wm_attributes("-topmost", True)  # Always on top

# Obține dimensiunile ecranului
screen_width = icon_root.winfo_screenwidth()
screen_height = icon_root.winfo_screenheight()

# Calculează coordonatele
x = screen_width - 30  # 30 este lățimea ferestrei pentru iconiță
y = screen_height // 2 - 15  # 15 este jumătate din înălțimea ferestrei pentru iconiță

icon_root.geometry(f'30x30+{x}+{y}')

icon_img = PhotoImage(file="pen.png")

icon_canvas = tk.Canvas(icon_root, width=30, height=30)
icon_canvas.pack()
icon_canvas.create_image(15, 15, image=icon_img)

icon_canvas.bind("<Button-1>", lambda e: toggle_main_window())
icon_root.mainloop()
root.mainloop()


