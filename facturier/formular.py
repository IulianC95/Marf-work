import tkinter as tk
from tkinter import ttk, simpledialog, Frame
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import json

# Variabile globale pentru a stoca produsele adăugate
added_products = []

# Lista predefinita de nume
name_list = ["George Screciu", "Cristi Domnea", "Alexandru Dinica", "Cuculici Iulian", "Marin Lucian", "Elena Gaina", "Cristina Tudor", "Lucian Tudor", "Marin Florin"]

def load_data(filename):
    try:
        with open(filename, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return []

def find_empty_row(sheet):
    for row in sheet.iter_rows(min_col=1, max_col=5):
        if all(cell.value is None for cell in row):
            return row[0].row
    return sheet.max_row + 1

def save_data(filename, data):
    with open(filename, 'w') as file:
        json.dump(data, file)

def add_new_item(item, filename):
    items = load_data(filename)
    if item not in items:
        items.append(item)
        save_data(filename, items)

def update_combobox_list(combobox, data_file):
    current_text = combobox.get()
    all_items = load_data(data_file)
    filtered_items = [item for item in all_items if current_text.lower() in item.lower()]
    combobox['values'] = filtered_items
    combobox.delete(0, tk.END)
    combobox.insert(0, current_text)

def add_product():
    global added_products
    product = product_combobox.get()
    quantity = quantity_entry.get()
    cost = cost_entry.get()
    total_cost = float(quantity) * float(cost)
    added_products.append(f"{product}(x{quantity} = {total_cost} RON)")
    product_combobox.set('+'.join(added_products))

def submit_data():
    global added_products

    selected_name = name_combobox.get()
    selected_date = calendar_entry.get_date()
    association_name = association_combobox.get()
    product_name = ' + '.join(added_products)
    total_sum = sum([float(product.split('= ')[1].split(' ')[0]) for product in added_products])

    add_new_item(association_name, 'associations.json')
    for product in added_products:
        add_new_item(product.split('(')[0].strip(), 'products.json')

    try:
        workbook = load_workbook(filename="\\\\192.168.0.99\\Administrare\\A S O C I A T I I\\Facturier.xlsx")
        sheet = workbook.active
        empty_row = find_empty_row(sheet)
        sheet.cell(row=empty_row, column=1, value=selected_name)
        sheet.cell(row=empty_row, column=2, value=selected_date)
        sheet.cell(row=empty_row, column=3, value=association_name)
        sheet.cell(row=empty_row, column=4, value=product_name)
        sheet.cell(row=empty_row, column=5, value=total_sum)
        workbook.save(filename="\\\\192.168.0.99\\Administrare\\A S O C I A T I I\\Facturier.xlsx")

        simpledialog.messagebox.showinfo("Informații introduse", f"Nume: {selected_name}\n"
                                    f"Data: {selected_date}\n"
                                    f"Asociație: {association_name}\n"
                                    f"Produse: {product_name}\n"
                                    f"Total: {total_sum} RON")
        added_products = []

    except PermissionError:
        simpledialog.messagebox.showerror("Eroare", "Fișierul Facturier.xlsx este deja deschis. Închideți fișierul și încercați din nou.")

def show_and_delete_products():
    def delete_selected_product():
        selected_product = product_listbox.get(product_listbox.curselection())
        products = load_data('products.json')
        if selected_product in products:
            products.remove(selected_product)
            save_data('products.json', products)
            product_listbox.delete(tk.ACTIVE)

    products_window = tk.Toplevel(app)
    products_window.title("Produse existente")
    products_window.geometry("300x450")

    product_listbox = tk.Listbox(products_window, width=40, height=20)
    product_listbox.pack(pady=20)

    for product in load_data('products.json'):
        product_listbox.insert(tk.END, product)

    delete_button = ttk.Button(products_window, text="Șterge produsul selectat", command=delete_selected_product)
    delete_button.pack(pady=10)

def show_and_delete_associations():
    def delete_selected_association():
        selected_association = association_listbox.get(association_listbox.curselection())
        associations = load_data('associations.json')
        if selected_association in associations:
            associations.remove(selected_association)
            save_data('associations.json', associations)
            association_listbox.delete(tk.ACTIVE)

    associations_window = tk.Toplevel(app)
    associations_window.title("Asociații existente")
    associations_window.geometry("300x450")

    association_listbox = tk.Listbox(associations_window, width=40, height=20)
    association_listbox.pack(pady=20)

    for association in load_data('associations.json'):
        association_listbox.insert(tk.END, association)

    delete_button = ttk.Button(associations_window, text="Șterge asociația selectată", command=delete_selected_association)
    delete_button.pack(pady=10)

# Functie noua de resetare a datelor
def reset_data():
    global added_products
    added_products = []
    name_combobox.set('')
    calendar_entry.set_date(None)
    association_combobox.set('')
    product_combobox.set('')
    quantity_entry.delete(0, tk.END)
    cost_entry.delete(0, tk.END)

app = tk.Tk()
app.title("MARF App")
app.geometry("800x600")
app.configure(bg="#2b66c4")

# Setează stilul pentru butoane
style = ttk.Style()

# Definirea stilului pentru butoane
style.configure("Custom.TButton",
                background="#5468ff",
                foreground="#000",
                font=("JetBrains Mono", 10),
                borderwidth=1,
                relief="raised",
                highlightthickness=0)
style.map("Custom.TButton",
          background=[("active", "#3c4fe0"), ("pressed", "#3c4fe0"), ("focus", "#3c4fe0")],
          relief=[("pressed", "sunken"), ("active", "raised")])

notebook = ttk.Notebook(app)
notebook.pack(pady=10, padx=10, expand=True, fill=tk.BOTH)



from fpdf import FPDF

def generate_pdf(date, name, report):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Titlu
    pdf.cell(200, 10, txt=f"Raport pentru data {date}", ln=True, align='C')

    # Numele
    pdf.cell(200, 10, txt=f"Raport realizat de: {name}", ln=True, align='L')

    # Raportul
    pdf.multi_cell(0, 10, txt=report)

    # Salvarea PDF-ului
    filename = f"Raport_{date.replace('-', '_')}_{name.replace(' ', '_')}.pdf"
    pdf.output(filename)
    simpledialog.messagebox.showinfo("PDF Generat", f"Raportul a fost salvat ca: {filename}")

def submit_daily_report():
    selected_date = daily_calendar_entry.get_date()
    selected_name = daily_name_combobox.get()
    report = daily_report_text.get("1.0", tk.END)

    generate_pdf(str(selected_date), selected_name, report)

# Crearea unui nou frame pentru tab-ul "Raport zilnic"
daily_report_frame = ttk.Frame(notebook)
notebook.add(daily_report_frame, text="Raport zilnic")

# Adăugarea unui calendar
daily_date_label = tk.Label(daily_report_frame, text="Selectează data:")
daily_date_label.pack(pady=10)
daily_calendar_entry = DateEntry(daily_report_frame, date_pattern='y-mm-dd', bg="white", fg="black")
daily_calendar_entry.pack(pady=10)

# Adăugarea unui dropdown pentru selecția numelui
daily_name_label = tk.Label(daily_report_frame, text="Selectează numele:")
daily_name_label.pack(pady=10)
daily_names = load_data('names.json')  # presupunem că avem un fișier 'names.json' cu numele
daily_name_combobox = ttk.Combobox(daily_report_frame, values=daily_names)
daily_name_combobox.pack(pady=10)

# Adăugarea unui text-area pentru raport
daily_report_label = tk.Label(daily_report_frame, text="Introduceți raportul:")
daily_report_label.pack(pady=10)
daily_report_text = tk.Text(daily_report_frame, width=70, height=15)
daily_report_text.pack(pady=10)

# Buton de trimitere
daily_submit_button = ttk.Button(daily_report_frame, text="Trimite raportul", command=submit_daily_report, style="Custom.TButton")
daily_submit_button.pack(pady=20)


formular_frame = ttk.Frame(notebook)
notebook.add(formular_frame, text="Facturier")

frame = tk.Frame(formular_frame, bg="#2b66c4", padx=20, pady=20)
frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

name_label = tk.Label(frame, text="Selectează numele:", bg="#2b66c4")
name_label.grid(row=0, column=0, sticky=tk.W, pady=5)
name_combobox = ttk.Combobox(frame, values=name_list, width=40)
name_combobox.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(0, 10))

date_label = tk.Label(frame, text="Selectează data:", bg="#2b66c4")
date_label.grid(row=1, column=0, sticky=tk.W, pady=5)
calendar_entry = DateEntry(frame, date_pattern='y-mm-dd', bg="white", fg="black")
calendar_entry.grid(row=1, column=1, sticky=tk.W, pady=5)

association_label = tk.Label(frame, text="Asociație:", bg="#2b66c4")
association_label.grid(row=2, column=0, sticky=tk.W, pady=5)
association_combobox = ttk.Combobox(frame, values=load_data('associations.json'), width=40)
association_combobox.grid(row=2, column=1, sticky=tk.W, pady=5, padx=(0, 10))
association_combobox.bind('<KeyRelease>', lambda event: update_combobox_list(association_combobox, 'associations.json'))

show_associations_button = ttk.Button(frame, text="Vizualizează Asociații", command=show_and_delete_associations, style="Custom.TButton")
show_associations_button.grid(row=2, column=2, sticky=tk.W, pady=5)
show_products_button = ttk.Button(frame, text="Vizualizează Produse", command=show_and_delete_products, style="Custom.TButton")
show_products_button.grid(row=1, column=2, sticky=tk.W, pady=5)


product_frame = Frame(frame, bd=2, relief="groove", bg="#2b66c4")
product_frame.grid(row=3, column=0, columnspan=3, pady=10, padx=5, sticky=tk.W+tk.E)

product_label = tk.Label(product_frame, text="Produs:", bg="#2b66c4")
product_label.grid(row=0, column=0, sticky=tk.W, pady=5)
product_combobox = ttk.Combobox(product_frame, values=load_data('products.json'), width=35)
product_combobox.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(0, 10))
product_combobox.bind('<KeyRelease>', lambda event: update_combobox_list(product_combobox, 'products.json'))


add_button = ttk.Button(product_frame, text="Adaugă", command=add_product, style="Custom.TButton")
add_button.grid(row=0, column=2, pady=5)


quantity_label = tk.Label(product_frame, text="Cantitate:", bg="#2b66c4")
quantity_label.grid(row=1, column=0, sticky=tk.W, pady=5)
quantity_entry = tk.Entry(product_frame)
quantity_entry.grid(row=1, column=1, sticky=tk.W, pady=5)

cost_label = tk.Label(product_frame, text="Cost per bucata:", bg="#2b66c4")
cost_label.grid(row=2, column=0, sticky=tk.W, pady=5)
cost_entry = tk.Entry(product_frame)
cost_entry.grid(row=2, column=1, sticky=tk.W, pady=5)

submit_button = ttk.Button(frame, text="Trimite", command=submit_data, style="Custom.TButton")
submit_button.grid(row=4, column=1, pady=10, sticky=tk.E + tk.W)

reset_button = ttk.Button(frame, text="Resetează", command=reset_data, style="Custom.TButton")
reset_button.grid(row=5, column=1, pady=10, sticky=tk.E + tk.W)

frame.grid_columnconfigure(0, weight=1)
frame.grid_columnconfigure(1, weight=1)

app.mainloop()
